#!/usr/bin/env python3
"""
Coffee Cue System Documentation Generator

This script generates comprehensive Excel documentation for the Coffee Cue System,
incorporating UI/UX information from screenshots with code structure details.
It creates a single Excel file with multiple worksheets covering all aspects of the system.

Requirements:
- openpyxl: pip install openpyxl
- Pillow: pip install Pillow (for screenshot processing, if needed)
"""

import os
import sys
import json
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import openpyxl.drawing.image

# Configuration
OUTPUT_FILE = "CoffeeCueSystemDocumentation.xlsx"
SCREENSHOTS_DIR = "./screenshots"  # Directory containing screenshots

# Color scheme
COLORS = {
    "header_bg": "4472C4",  # Blue header background
    "header_text": "FFFFFF",  # White header text
    "alt_row": "E6F0FF",  # Light blue alternating row
    "highlight": "FFEB9C",  # Yellow highlight
    "success": "C6EFCE",  # Green success
    "warning": "FFEB9C",  # Yellow warning
    "error": "FFC7CE",  # Red error
    "link": "0563C1",  # Blue hyperlink
    "section_bg": {
        "frontend": "DDEBF7",  # Light blue for frontend
        "backend": "E2EFDA",   # Light green for backend
        "database": "FCE4D6",  # Light orange for database
        "auth": "E1D9F2",      # Light purple for auth
        "api": "FFF2CC",       # Light yellow for API
    }
}

def setup_worksheet(wb, title, headers, col_widths=None):
    """Create and format a worksheet with headers."""
    ws = wb.create_sheet(title)
    
    # Set column widths if provided
    if col_widths:
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    # Add headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True, color=COLORS["header_text"])
        cell.fill = PatternFill(start_color=COLORS["header_bg"], end_color=COLORS["header_bg"], fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Add thin borders to header cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_idx in range(1, len(headers) + 1):
        ws.cell(row=1, column=col_idx).border = thin_border
    
    return ws

def load_existing_documentation():
    """
    Load data from an existing documentation file if it exists.
    Returns a dictionary with worksheet names as keys and row data as values.
    """
    existing_data = {}
    
    if os.path.exists(OUTPUT_FILE):
        print(f"Loading existing documentation from {OUTPUT_FILE}")
        try:
            wb = load_workbook(OUTPUT_FILE)
            for sheet_name in wb.sheetnames:
                if sheet_name != "README":  # Skip the README sheet
                    ws = wb[sheet_name]
                    headers = [cell.value for cell in ws[1]]
                    
                    rows = []
                    for row_idx in range(2, ws.max_row + 1):
                        row_data = [ws.cell(row=row_idx, column=col_idx).value 
                                   for col_idx in range(1, len(headers) + 1)]
                        rows.append(row_data)
                    
                    existing_data[sheet_name] = {
                        "headers": headers,
                        "rows": rows
                    }
            print(f"Loaded {len(existing_data)} worksheets from existing documentation")
        except Exception as e:
            print(f"Error loading existing documentation: {e}")
            print("Creating new documentation file")
    
    return existing_data

def populate_file_components(wb, existing_data=None):
    """Populate the File Components worksheet."""
    headers = [
        "File Path", "File Name", "Purpose/Description", "Type", 
        "Dependencies", "Related UI Components", "Last Modified"
    ]
    col_widths = [30, 25, 60, 15, 40, 40, 20]
    
    if existing_data and "File Components" in existing_data:
        rows = existing_data["File Components"]["rows"]
        # Verify headers match
        if existing_data["File Components"]["headers"] == headers:
            ws = setup_worksheet(wb, "File Components", headers, col_widths)
            
            # Populate with existing data
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    
            print(f"Populated File Components with {len(rows)} existing records")
            return ws
    
    # Create new worksheet if no existing data or headers don't match
    ws = setup_worksheet(wb, "File Components", headers, col_widths)
    
    # Add data from the project files
    files = [
        {
            "path": "src/components",
            "name": "BaristaInterface.js",
            "purpose": "Main barista dashboard UI component that displays pending orders, current order, and completed orders",
            "type": "React Component",
            "dependencies": "useOrders, OrderDataService, MessageService, AuthService",
            "ui_components": "HeaderBar, NavTabs, OrdersTab, CompletedTab, DisplayTab",
            "last_modified": "2025-03-01"
        },
        {
            "path": "src/services",
            "name": "AuthService.js",
            "purpose": "Handles authentication with JWT tokens, login/logout functionality, and session management",
            "type": "Service",
            "dependencies": "OrderDataService",
            "ui_components": "LoginPage, ProtectedRoute",
            "last_modified": "2025-02-15"
        },
        {
            "path": "src/services",
            "name": "OrderDataService.js",
            "purpose": "Manages all API calls related to coffee orders, including fetching, creating, and updating orders",
            "type": "Service",
            "dependencies": "None",
            "ui_components": "PendingOrdersSection, InProgressOrdersList, CompletedOrderCard",
            "last_modified": "2025-03-05"
        },
        {
            "path": "src/hooks",
            "name": "useOrders.js",
            "purpose": "Custom React hook for managing order state and operations throughout the application",
            "type": "React Hook",
            "dependencies": "OrderDataService, useState, useEffect, useCallback",
            "ui_components": "BaristaInterface, PendingOrdersSection",
            "last_modified": "2025-03-04"
        },
        {
            "path": "src/components",
            "name": "PendingOrdersSection.js",
            "purpose": "Displays pending orders with filtering and batch processing capabilities",
            "type": "React Component",
            "dependencies": "VipOrdersList, BatchOrdersList, RegularOrdersList",
            "ui_components": "Seen in Image 9, shows order cards in upcoming orders section",
            "last_modified": "2025-03-02"
        },
        {
            "path": "backend/routes",
            "name": "api_routes.py",
            "purpose": "Defines API endpoints for order management, including CRUD operations for orders",
            "type": "Flask Routes",
            "dependencies": "Flask, models.orders, auth",
            "ui_components": "Indirectly used by all order management UI components",
            "last_modified": "2025-02-28"
        },
        {
            "path": "backend/routes",
            "name": "auth_routes.py",
            "purpose": "Defines API endpoints for authentication, login, token refresh",
            "type": "Flask Routes",
            "dependencies": "Flask, models.users, jwt_extended",
            "ui_components": "Login screen, role selection (Image 18)",
            "last_modified": "2025-02-20"
        },
        {
            "path": "backend/services",
            "name": "messaging.py",
            "purpose": "Handles SMS notification integration with Twilio for order status updates",
            "type": "Service Class",
            "dependencies": "Twilio, models.orders",
            "ui_components": "MessageDialog, SMS Configuration in settings",
            "last_modified": "2025-02-25"
        },
        {
            "path": "backend/models",
            "name": "orders.py",
            "purpose": "Database models and business logic for order management",
            "type": "Model",
            "dependencies": "PostgreSQL",
            "ui_components": "All order-related UI components",
            "last_modified": "2025-03-01"
        },
        {
            "path": "backend/models",
            "name": "stations.py",
            "purpose": "Database models and business logic for coffee station management",
            "type": "Model",
            "dependencies": "PostgreSQL",
            "ui_components": "Station Management screens (Image 4, Image 14)",
            "last_modified": "2025-02-22"
        }
    ]
    
    for row_idx, file_info in enumerate(files, 2):
        ws.cell(row=row_idx, column=1, value=file_info["path"])
        ws.cell(row=row_idx, column=2, value=file_info["name"])
        ws.cell(row=row_idx, column=3, value=file_info["purpose"])
        ws.cell(row=row_idx, column=4, value=file_info["type"])
        ws.cell(row=row_idx, column=5, value=file_info["dependencies"])
        ws.cell(row=row_idx, column=6, value=file_info["ui_components"])
        ws.cell(row=row_idx, column=7, value=file_info["last_modified"])
    
    # Apply alternating row colors
    for row_idx in range(2, len(files) + 2):
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color=COLORS["alt_row"],
                    end_color=COLORS["alt_row"],
                    fill_type="solid"
                )
    
    print(f"Created File Components worksheet with {len(files)} entries")
    return ws

def populate_frontend_architecture(wb, existing_data=None):
    """Populate the Frontend Architecture worksheet."""
    headers = [
        "Component Name", "Parent Component", "Function/Class", "Props/Parameters", 
        "State Management", "API Endpoints", "Related Files", "UI Screenshot Reference"
    ]
    col_widths = [25, 25, 15, 40, 40, 40, 30, 25]
    
    if existing_data and "Frontend Architecture" in existing_data:
        rows = existing_data["Frontend Architecture"]["rows"]
        # Verify headers match
        if existing_data["Frontend Architecture"]["headers"] == headers:
            ws = setup_worksheet(wb, "Frontend Architecture", headers, col_widths)
            
            # Populate with existing data
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    
            print(f"Populated Frontend Architecture with {len(rows)} existing records")
            return ws
    
    # Create new worksheet if no existing data or headers don't match
    ws = setup_worksheet(wb, "Frontend Architecture", headers, col_widths)
    
    # Add components from the diagram and screenshots
    components = [
        {
            "name": "BaristaInterface",
            "parent": "App",
            "type": "Function Component",
            "props": "None",
            "state": "useState for activeTab, selectedOrderId, batchModeActive, waitTime, stations, selectedStation",
            "endpoints": "/api/orders/pending, /api/orders/in-progress, /api/orders/completed",
            "related_files": "useOrders.js, OrderDataService.js, MessageService.js",
            "screenshot": "Image 1 (Component Structure), Image 9 (Implementation)"
        },
        {
            "name": "HeaderBar",
            "parent": "BaristaInterface",
            "type": "Function Component",
            "props": "stationName, online, queueCount, waitTime, onHelp",
            "state": "useState for showStationSelector",
            "endpoints": "None",
            "related_files": "None",
            "screenshot": "Image 9 (Top brown bar)"
        },
        {
            "name": "NavTabs",
            "parent": "BaristaInterface",
            "type": "Function Component",
            "props": "activeTab, setActiveTab",
            "state": "None",
            "endpoints": "None",
            "related_files": "None",
            "screenshot": "Image 9 (Orders, Stock, Schedule, etc. tabs)"
        },
        {
            "name": "OrdersTab",
            "parent": "BaristaInterface",
            "type": "Function Component",
            "props": "pendingOrders, inProgressOrders, onStartOrder, onCompleteOrder, onProcessBatch",
            "state": "None",
            "endpoints": "None",
            "related_files": "CurrentOrderSection.js, PendingOrdersSection.js",
            "screenshot": "Image 9 (Main content area)"
        },
        {
            "name": "CurrentOrderSection",
            "parent": "OrdersTab",
            "type": "Function Component",
            "props": "currentOrder, onCompleteOrder, onSendMessage, onPrintLabel",
            "state": "None",
            "endpoints": "/api/orders/{id}/complete",
            "related_files": "OrderDataService.js",
            "screenshot": "Image 9 (Left side with Michael Johnson order)"
        },
        {
            "name": "PendingOrdersSection",
            "parent": "OrdersTab",
            "type": "Function Component",
            "props": "orders, filter, onFilterChange, onStartOrder, onProcessBatch",
            "state": "None",
            "endpoints": "/api/orders/{id}/start, /api/orders/batch",
            "related_files": "VipOrdersList.js, BatchGroupsList.js, RegularOrdersList.js",
            "screenshot": "Image 9 (Right side with upcoming orders)"
        },
        {
            "name": "BatchGroupsList",
            "parent": "PendingOrdersSection",
            "type": "Function Component",
            "props": "batchGroups, onStartOrder, onProcessBatch",
            "state": "None",
            "endpoints": "None",
            "related_files": "None",
            "screenshot": "Image 9 (Batch: Latte Soy section)"
        },
        {
            "name": "WaitTimeDialog",
            "parent": "BaristaInterface",
            "type": "Function Component",
            "props": "currentWaitTime, onSubmit, onClose",
            "state": "useState for new wait time",
            "endpoints": "/api/settings/wait-time",
            "related_files": "None",
            "screenshot": "Not directly visible in screenshots"
        },
        {
            "name": "WalkInOrderDialog",
            "parent": "BaristaInterface",
            "type": "Function Component",
            "props": "onSubmit, onClose",
            "state": "useState for order form fields",
            "endpoints": "/api/orders/walk-in",
            "related_files": "None",
            "screenshot": "Not directly visible in screenshots"
        },
        {
            "name": "MessageDialog",
            "parent": "BaristaInterface",
            "type": "Function Component",
            "props": "order, onSubmit, onClose",
            "state": "useState for message text",
            "endpoints": "/api/orders/{id}/message",
            "related_files": "MessageService.js",
            "screenshot": "Not directly visible in screenshots"
        },
        {
            "name": "DisplayScreen",
            "parent": "App",
            "type": "Function Component",
            "props": "None",
            "state": "useState for completedOrders, inProgressOrders",
            "endpoints": "/api/orders/in-progress, /api/orders/completed",
            "related_files": "OrderDataService.js",
            "screenshot": "Image 2 (Customer-facing display)"
        },
        {
            "name": "StationManagement",
            "parent": "App",
            "type": "Function Component",
            "props": "None",
            "state": "useState for stations, inventory",
            "endpoints": "/api/stations, /api/inventory",
            "related_files": "StationService.js",
            "screenshot": "Image 4 (Station Management), Image 14 (Barista Stations Management)"
        },
        {
            "name": "SystemSettings",
            "parent": "App",
            "type": "Function Component",
            "props": "None",
            "state": "useState for settings fields",
            "endpoints": "/api/settings",
            "related_files": "SettingsService.js",
            "screenshot": "Image 3 (System Settings), Image 11 (System Settings)"
        },
        {
            "name": "SystemMonitoring",
            "parent": "App",
            "type": "Function Component",
            "props": "None",
            "state": "useState for serviceStatus, logs",
            "endpoints": "/api/monitoring, /api/logs",
            "related_files": "MonitoringService.js",
            "screenshot": "Image 6-7 (System Monitoring)"
        },
        {
            "name": "SupportDashboard",
            "parent": "App",
            "type": "Function Component",
            "props": "None",
            "state": "useState for stats, alerts, recentLogs",
            "endpoints": "/api/stats, /api/alerts, /api/logs",
            "related_files": "None",
            "screenshot": "Image 8 (Support Dashboard)"
        }
    ]
    
    for row_idx, component in enumerate(components, 2):
        ws.cell(row=row_idx, column=1, value=component["name"])
        ws.cell(row=row_idx, column=2, value=component["parent"])
        ws.cell(row=row_idx, column=3, value=component["type"])
        ws.cell(row=row_idx, column=4, value=component["props"])
        ws.cell(row=row_idx, column=5, value=component["state"])
        ws.cell(row=row_idx, column=6, value=component["endpoints"])
        ws.cell(row=row_idx, column=7, value=component["related_files"])
        ws.cell(row=row_idx, column=8, value=component["screenshot"])
    
    # Apply alternating row colors
    for row_idx in range(2, len(components) + 2):
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color=COLORS["alt_row"],
                    end_color=COLORS["alt_row"],
                    fill_type="solid"
                )
    
    print(f"Created Frontend Architecture worksheet with {len(components)} entries")
    return ws

def populate_backend_architecture(wb, existing_data=None):
    """Populate the Backend Architecture worksheet."""
    headers = [
        "Route Name", "HTTP Method", "Endpoint Path", "Controller/Handler", 
        "Required Auth Role", "Request Parameters", "Response Structure", "UI Integration"
    ]
    col_widths = [25, 12, 30, 20, 20, 40, 40, 30]
    
    if existing_data and "Backend Architecture" in existing_data:
        rows = existing_data["Backend Architecture"]["rows"]
        # Verify headers match
        if existing_data["Backend Architecture"]["headers"] == headers:
            ws = setup_worksheet(wb, "Backend Architecture", headers, col_widths)
            
            # Populate with existing data
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    
            print(f"Populated Backend Architecture with {len(rows)} existing records")
            return ws
    
    # Create new worksheet if no existing data or headers don't match
    ws = setup_worksheet(wb, "Backend Architecture", headers, col_widths)
    
    # Add routes and endpoints
    routes = [
        {
            "name": "get_pending_orders",
            "method": "GET",
            "path": "/api/orders/pending",
            "controller": "api_routes.py",
            "auth_role": "barista, admin",
            "params": "None",
            "response": "JSON array of pending order objects",
            "ui": "PendingOrdersSection in BaristaInterface (Image 9)"
        },
        {
            "name": "get_in_progress_orders",
            "method": "GET",
            "path": "/api/orders/in-progress",
            "controller": "api_routes.py",
            "auth_role": "barista, admin",
            "params": "None",
            "response": "JSON array of in-progress order objects",
            "ui": "CurrentOrderSection in BaristaInterface (Image 9)"
        },
        {
            "name": "get_completed_orders",
            "method": "GET",
            "path": "/api/orders/completed",
            "controller": "api_routes.py",
            "auth_role": "barista, admin",
            "params": "None",
            "response": "JSON array of completed order objects",
            "ui": "CompletedOrderCard in CompletedTab"
        },
        {
            "name": "start_order",
            "method": "POST",
            "path": "/api/orders/{order_id}/start",
            "controller": "api_routes.py",
            "auth_role": "barista, admin",
            "params": "order_id (URL parameter)",
            "response": "JSON success status",
            "ui": "Start buttons in PendingOrdersSection (Image 9)"
        },
        {
            "name": "complete_order",
            "method": "POST",
            "path": "/api/orders/{order_id}/complete",
            "controller": "api_routes.py",
            "auth_role": "barista, admin",
            "params": "order_id (URL parameter)",
            "response": "JSON success status",
            "ui": "COMPLETE ORDER button in CurrentOrderSection (Image 9)"
        },
        {
            "name": "batch_process_orders",
            "method": "POST",
            "path": "/api/orders/batch",
            "controller": "api_routes.py",
            "auth_role": "barista, admin",
            "params": "order_ids (array), action (start/complete)",
            "response": "JSON success status with processed count",
            "ui": "Process Batch button in BatchGroupsList (Image 9)"
        },
        {
            "name": "send_message_to_customer",
            "method": "POST",
            "path": "/api/orders/{order_id}/message",
            "controller": "sms_routes.py",
            "auth_role": "barista, admin",
            "params": "order_id (URL parameter), message (body)",
            "response": "JSON success status",
            "ui": "Message button and MessageDialog in BaristaInterface"
        },
        {
            "name": "update_wait_time",
            "method": "POST",
            "path": "/api/settings/wait-time",
            "controller": "settings_routes.py",
            "auth_role": "barista, admin",
            "params": "waitTime (integer)",
            "response": "JSON success status",
            "ui": "WaitTimeDialog accessed from ActionBar (Image 9)"
        },
        {
            "name": "add_walk_in_order",
            "method": "POST",
            "path": "/api/orders/walk-in",
            "controller": "api_routes.py",
            "auth_role": "barista, admin",
            "params": "Customer details and order details in request body",
            "response": "JSON with order details including ID",
            "ui": "WalkInOrderDialog accessed from ActionBar (Image 9)"
        },
        {
            "name": "login",
            "method": "POST",
            "path": "/api/auth/login",
            "controller": "auth_routes.py",
            "auth_role": "None",
            "params": "username, password",
            "response": "JWT token, refresh token, user object",
            "ui": "Login screen (not shown in screenshots)"
        },
        {
            "name": "refresh_token",
            "method": "POST",
            "path": "/api/auth/refresh",
            "controller": "auth_routes.py",
            "auth_role": "None",
            "params": "refreshToken",
            "response": "New JWT token",
            "ui": "Handled automatically by AuthService"
        },
        {
            "name": "get_stations",
            "method": "GET",
            "path": "/api/stations",
            "controller": "station_routes.py",
            "auth_role": "barista, admin",
            "params": "None",
            "response": "JSON array of station objects",
            "ui": "StationManagement (Image 4, Image 14)"
        },
        {
            "name": "get_inventory",
            "method": "GET",
            "path": "/api/inventory",
            "controller": "inventory_routes.py",
            "auth_role": "barista, admin",
            "params": "None",
            "response": "JSON inventory data by station",
            "ui": "Station Inventory Status (Image 14)"
        },
        {
            "name": "update_inventory",
            "method": "POST",
            "path": "/api/inventory/update",
            "controller": "inventory_routes.py",
            "auth_role": "barista, admin",
            "params": "station_id, item_type, amount",
            "response": "JSON success status",
            "ui": "Restock buttons in Station Inventory Status (Image 14)"
        }
    ]
    
    for row_idx, route in enumerate(routes, 2):
        ws.cell(row=row_idx, column=1, value=route["name"])
        ws.cell(row=row_idx, column=2, value=route["method"])
        ws.cell(row=row_idx, column=3, value=route["path"])
        ws.cell(row=row_idx, column=4, value=route["controller"])
        ws.cell(row=row_idx, column=5, value=route["auth_role"])
        ws.cell(row=row_idx, column=6, value=route["params"])
        ws.cell(row=row_idx, column=7, value=route["response"])
        ws.cell(row=row_idx, column=8, value=route["ui"])
    
    # Apply alternating row colors
    for row_idx in range(2, len(routes) + 2):
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color=COLORS["alt_row"],
                    end_color=COLORS["alt_row"],
                    fill_type="solid"
                )
    
    print(f"Created Backend Architecture worksheet with {len(routes)} entries")
    return ws

def populate_database_schema(wb, existing_data=None):
    """Populate the Database Schema worksheet."""
    headers = [
        "Table Name", "Column Name", "Data Type", "Constraints", 
        "References", "Purpose", "Related Models", "UI Representation"
    ]
    col_widths = [20, 25, 15, 20, 20, 40, 20, 30]
    
    if existing_data and "Database Schema" in existing_data:
        rows = existing_data["Database Schema"]["rows"]
        # Verify headers match
        if existing_data["Database Schema"]["headers"] == headers:
            ws = setup_worksheet(wb, "Database Schema", headers, col_widths)
            
            # Populate with existing data
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    
            print(f"Populated Database Schema with {len(rows)} existing records")
            return ws
    
    # Create new worksheet if no existing data or headers don't match
    ws = setup_worksheet(wb, "Database Schema", headers, col_widths)
    
    # Add schema information
    schema = [
        # orders table
        {
            "table": "orders",
            "column": "id",
            "type": "SERIAL",
            "constraints": "PRIMARY KEY",
            "references": "",
            "purpose": "Unique identifier for orders",
            "models": "models/orders.py",
            "ui": "Order IDs displayed in all order cards (eg. A082534)"
        },
        {
            "table": "orders",
            "column": "order_number",
            "type": "VARCHAR(20)",
            "constraints": "UNIQUE NOT NULL",
            "references": "",
            "purpose": "Human-readable order identifier",
            "models": "models/orders.py",
            "ui": "Order numbers displayed in all order cards"
        },
        {
            "table": "orders",
            "column": "phone",
            "type": "VARCHAR(20)",
            "constraints": "NOT NULL",
            "references": "",
            "purpose": "Customer phone number for SMS notifications",
            "models": "models/orders.py",
            "ui": "Shown as customer contact in order cards"
        },
        {
            "table": "orders",
            "column": "order_details",
            "type": "JSONB",
            "constraints": "NOT NULL",
            "references": "",
            "purpose": "Details of the coffee order (type, milk, sugar, etc.)",
            "models": "models/orders.py",
            "ui": "Displayed in order cards (Image 9: Large Cappuccino, Oat milk, etc.)"
        },
        {
            "table": "orders",
            "column": "status",
            "type": "VARCHAR(20)",
            "constraints": "NOT NULL",
            "references": "",
            "purpose": "Current order status (pending, in-progress, completed)",
            "models": "models/orders.py",
            "ui": "Determines which section order appears in (Image 2, Image 9)"
        },
        {
            "table": "orders",
            "column": "station_id",
            "type": "INTEGER",
            "constraints": "NOT NULL",
            "references": "",
            "purpose": "Station assigned to prepare the order",
            "models": "models/orders.py",
            "ui": "Shown in station header (Image 2, Image 9)"
        },
        {
            "table": "orders",
            "column": "created_at",
            "type": "TIMESTAMP",
            "constraints": "DEFAULT CURRENT_TIMESTAMP",
            "references": "",
            "purpose": "When the order was created",
            "models": "models/orders.py",
            "ui": "Used to calculate wait times shown in UI"
        },
        {
            "table": "orders",
            "column": "queue_priority",
            "type": "INTEGER",
            "constraints": "NOT NULL DEFAULT 5",
            "references": "",
            "purpose": "Priority level for sorting (1=highest)",
            "models": "models/orders.py",
            "ui": "VIP orders shown at top of queue (Image 9)"
        },
        
        # station_stats table
        {
            "table": "station_stats",
            "column": "station_id",
            "type": "INTEGER",
            "constraints": "PRIMARY KEY",
            "references": "",
            "purpose": "Unique identifier for stations",
            "models": "models/stations.py",
            "ui": "Station cards in station management (Image 14)"
        },
        {
            "table": "station_stats",
            "column": "current_load",
            "type": "INTEGER",
            "constraints": "DEFAULT 0",
            "references": "",
            "purpose": "Number of orders currently assigned",
            "models": "models/stations.py",
            "ui": "Queue count in station cards (Image 14)"
        },
        {
            "table": "station_stats",
            "column": "avg_completion_time",
            "type": "INTEGER",
            "constraints": "DEFAULT 180",
            "references": "",
            "purpose": "Average time to complete orders in seconds",
            "models": "models/stations.py",
            "ui": "Used to calculate wait times"
        },
        {
            "table": "station_stats",
            "column": "status",
            "type": "VARCHAR(20)",
            "constraints": "DEFAULT 'active'",
            "references": "",
            "purpose": "Station status (active, inactive, maintenance)",
            "models": "models/stations.py",
            "ui": "Status indicator in Station Management (Image 4, Image 14)"
        },
        {
            "table": "station_stats",
            "column": "barista_name",
            "type": "VARCHAR(100)",
            "constraints": "",
            "references": "",
            "purpose": "Name of assigned barista",
            "models": "models/stations.py",
            "ui": "Barista name in station cards (Image 14)"
        },
        
        # users table
        {
            "table": "users",
            "column": "id",
            "type": "SERIAL",
            "constraints": "PRIMARY KEY",
            "references": "",
            "purpose": "Unique identifier for users",
            "models": "models/users.py",
            "ui": "Not directly visible, used in auth"
        },
        {
            "table": "users",
            "column": "username",
            "type": "VARCHAR(50)",
            "constraints": "UNIQUE NOT NULL",
            "references": "",
            "purpose": "Login username",
            "models": "models/users.py",
            "ui": "Login screen (not shown in screenshots)"
        },
        {
            "table": "users",
            "column": "password_hash",
            "type": "TEXT",
            "constraints": "NOT NULL",
            "references": "",
            "purpose": "Hashed password for authentication",
            "models": "models/users.py",
            "ui": "Login screen (not shown in screenshots)"
        },
        {
            "table": "users",
            "column": "role",
            "type": "VARCHAR(20)",
            "constraints": "NOT NULL",
            "references": "",
            "purpose": "User role (admin, barista, staff, support, display)",
            "models": "models/users.py",
            "ui": "Role selection screen (Image 18)"
        },
        
        # settings table
        {
            "table": "settings",
            "column": "key",
            "type": "VARCHAR(100)",
            "constraints": "PRIMARY KEY",
            "references": "",
            "purpose": "Setting identifier",
            "models": "models/settings.py",
            "ui": "System Settings screens (Image 3, Image 10, Image 11)"
        },
        {
            "table": "settings",
            "column": "value",
            "type": "TEXT",
            "constraints": "NOT NULL",
            "references": "",
            "purpose": "Setting value",
            "models": "models/settings.py",
            "ui": "Input fields in Settings screens"
        },
    ]
    
    for row_idx, column in enumerate(schema, 2):
        ws.cell(row=row_idx, column=1, value=column["table"])
        ws.cell(row=row_idx, column=2, value=column["column"])
        ws.cell(row=row_idx, column=3, value=column["type"])
        ws.cell(row=row_idx, column=4, value=column["constraints"])
        ws.cell(row=row_idx, column=5, value=column["references"])
        ws.cell(row=row_idx, column=6, value=column["purpose"])
        ws.cell(row=row_idx, column=7, value=column["models"])
        ws.cell(row=row_idx, column=8, value=column["ui"])
    
    # Format the table name to span multiple rows
    current_table = None
    start_row = 2
    
    for row_idx in range(2, len(schema) + 2):
        table_name = ws.cell(row=row_idx, column=1).value
        
        # Apply alternating row colors
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color=COLORS["alt_row"],
                    end_color=COLORS["alt_row"],
                    fill_type="solid"
                )
        
        # If we've moved to a new table, add table highlighting
        if table_name != current_table:
            if current_table is not None:
                # Highlight the first cell of each table section
                for r in range(start_row, row_idx):
                    ws.cell(row=r, column=1).fill = PatternFill(
                        start_color=COLORS["section_bg"]["database"],
                        end_color=COLORS["section_bg"]["database"],
                        fill_type="solid"
                    )
                    ws.cell(row=r, column=1).font = Font(bold=(r == start_row))
            
            current_table = table_name
            start_row = row_idx
    
    # Handle the last table
    if current_table is not None:
        for r in range(start_row, len(schema) + 2):
            ws.cell(row=r, column=1).fill = PatternFill(
                start_color=COLORS["section_bg"]["database"],
                end_color=COLORS["section_bg"]["database"],
                fill_type="solid"
            )
            ws.cell(row=r, column=1).font = Font(bold=(r == start_row))
    
    print(f"Created Database Schema worksheet with {len(schema)} entries")
    return ws

def populate_api_endpoints(wb, existing_data=None):
    """Populate the API Endpoints worksheet."""
    headers = [
        "Path", "Method", "Description", "Request Format", 
        "Response Format", "Auth Required", "Used By", "Screenshot Reference"
    ]
    col_widths = [30, 10, 40, 40, 40, 15, 30, 20]
    
    if existing_data and "API Endpoints" in existing_data:
        rows = existing_data["API Endpoints"]["rows"]
        # Verify headers match
        if existing_data["API Endpoints"]["headers"] == headers:
            ws = setup_worksheet(wb, "API Endpoints", headers, col_widths)
            
            # Populate with existing data
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    
            print(f"Populated API Endpoints with {len(rows)} existing records")
            return ws
    
    # Create new worksheet if no existing data or headers don't match
    ws = setup_worksheet(wb, "API Endpoints", headers, col_widths)
    
    # Add API endpoint information
    endpoints = [
        {
            "path": "/api/orders/pending",
            "method": "GET",
            "description": "Get all pending orders",
            "request": "None",
            "response": "Array of order objects with status 'pending'",
            "auth": "Yes - barista",
            "used_by": "PendingOrdersSection component",
            "screenshot": "Used in Image 9 for Upcoming Orders"
        },
        {
            "path": "/api/orders/in-progress",
            "method": "GET",
            "description": "Get all in-progress orders",
            "request": "None",
            "response": "Array of order objects with status 'in-progress'",
            "auth": "Yes - barista",
            "used_by": "CurrentOrderSection component",
            "screenshot": "Used in Image 9 for Current Order"
        },
        {
            "path": "/api/orders/completed",
            "method": "GET",
            "description": "Get all completed orders",
            "request": "None",
            "response": "Array of order objects with status 'completed'",
            "auth": "Yes - barista",
            "used_by": "CompletedTab component",
            "screenshot": "Used in customer display (Image 2)"
        },
        {
            "path": "/api/orders/{id}/start",
            "method": "POST",
            "description": "Start processing an order",
            "request": "None (order ID in URL)",
            "response": "{'success': true}",
            "auth": "Yes - barista",
            "used_by": "Start buttons in PendingOrdersSection",
            "screenshot": "Start buttons in Image 9"
        },
        {
            "path": "/api/orders/{id}/complete",
            "method": "POST",
            "description": "Mark an order as completed",
            "request": "None (order ID in URL)",
            "response": "{'success': true}",
            "auth": "Yes - barista",
            "used_by": "COMPLETE ORDER button in CurrentOrderSection",
            "screenshot": "Complete button in Image 9"
        },
        {
            "path": "/api/orders/batch",
            "method": "POST",
            "description": "Process multiple orders at once",
            "request": "{'order_ids': [ids], 'action': 'start'/'complete'}",
            "response": "{'success': true, 'processed': count}",
            "auth": "Yes - barista",
            "used_by": "Process Batch button in BatchGroupsList",
            "screenshot": "Batch processing in Image 9"
        },
        {
            "path": "/api/orders/walk-in",
            "method": "POST",
            "description": "Create a new walk-in order",
            "request": "Order details object",
            "response": "Created order object with ID",
            "auth": "Yes - barista",
            "used_by": "Add Walk-in Order button in ActionBar",
            "screenshot": "'Add Walk-in Order' button in Image 9"
        },
        {
            "path": "/api/orders/{id}/message",
            "method": "POST",
            "description": "Send SMS message to customer",
            "request": "{'message': 'text'}",
            "response": "{'success': true}",
            "auth": "Yes - barista",
            "used_by": "Message buttons and MessageDialog",
            "screenshot": "Message buttons in Image 9"
        },
        {
            "path": "/api/stations",
            "method": "GET",
            "description": "Get all coffee stations",
            "request": "None",
            "response": "Array of station objects",
            "auth": "Yes - barista",
            "used_by": "Station selector in HeaderBar",
            "screenshot": "Station selection in Image 9, full view in Image 14"
        },
        {
            "path": "/api/stations/{id}",
            "method": "GET",
            "description": "Get single station details",
            "request": "None (station ID in URL)",
            "response": "Station object with details",
            "auth": "Yes - barista",
            "used_by": "StationManagement component",
            "screenshot": "Station details in Image 14"
        },
        {
            "path": "/api/inventory",
            "method": "GET",
            "description": "Get inventory status for all stations",
            "request": "None",
            "response": "Inventory levels object by station and item type",
            "auth": "Yes - barista",
            "used_by": "StockTab component",
            "screenshot": "Inventory display in Image 14"
        },
        {
            "path": "/api/settings/wait-time",
            "method": "POST",
            "description": "Update default wait time",
            "request": "{'waitTime': minutes}",
            "response": "{'success': true}",
            "auth": "Yes - barista",
            "used_by": "WaitTimeDialog component",
            "screenshot": "Wait time displayed in Image 9 header"
        },
        {
            "path": "/api/auth/login",
            "method": "POST",
            "description": "User login",
            "request": "{'username': 'user', 'password': 'pass'}",
            "response": "{'token': 'jwt', 'refreshToken': 'token', 'user': {}}",
            "auth": "No",
            "used_by": "AuthService.login()",
            "screenshot": "Role selection screen (Image 18)"
        },
        {
            "path": "/api/auth/refresh",
            "method": "POST",
            "description": "Refresh JWT token",
            "request": "{'refreshToken': 'token'}",
            "response": "{'token': 'new-jwt'}",
            "auth": "No",
            "used_by": "AuthService.refreshToken()",
            "screenshot": "N/A"
        },
        {
            "path": "/api/chat/messages",
            "method": "GET",
            "description": "Get chat messages for barista stations",
            "request": "None",
            "response": "Array of chat message objects",
            "auth": "Yes - barista",
            "used_by": "StationChat component",
            "screenshot": "N/A"
        },
        {
            "path": "/api/monitoring",
            "method": "GET",
            "description": "Get system monitoring status",
            "request": "None",
            "response": "System status object (server, database, services)",
            "auth": "Yes - support",
            "used_by": "SystemMonitoring component",
            "screenshot": "System Monitoring (Images 6-7)"
        },
        {
            "path": "/api/logs",
            "method": "GET",
            "description": "Get system logs",
            "request": "Optional filter parameters",
            "response": "Array of log entries",
            "auth": "Yes - support",
            "used_by": "SystemLogs component",
            "screenshot": "System Logs (Image 5)"
        }
    ]
    
    for row_idx, endpoint in enumerate(endpoints, 2):
        ws.cell(row=row_idx, column=1, value=endpoint["path"])
        ws.cell(row=row_idx, column=2, value=endpoint["method"])
        ws.cell(row=row_idx, column=3, value=endpoint["description"])
        ws.cell(row=row_idx, column=4, value=endpoint["request"])
        ws.cell(row=row_idx, column=5, value=endpoint["response"])
        ws.cell(row=row_idx, column=6, value=endpoint["auth"])
        ws.cell(row=row_idx, column=7, value=endpoint["used_by"])
        ws.cell(row=row_idx, column=8, value=endpoint["screenshot"])
    
    # Apply alternating row colors
    for row_idx in range(2, len(endpoints) + 2):
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color=COLORS["alt_row"],
                    end_color=COLORS["alt_row"],
                    fill_type="solid"
                )
    
    print(f"Created API Endpoints worksheet with {len(endpoints)} entries")
    return ws

def populate_user_workflows(wb, existing_data=None):
    """Populate the User Workflows worksheet."""
    headers = [
        "Workflow Name", "Actor", "Step Number", "Action", 
        "Frontend Component", "Backend Route", "Notes", "Screenshot Reference"
    ]
    col_widths = [25, 15, 10, 40, 25, 25, 40, 20]
    
    if existing_data and "User Workflows" in existing_data:
        rows = existing_data["User Workflows"]["rows"]
        # Verify headers match
        if existing_data["User Workflows"]["headers"] == headers:
            ws = setup_worksheet(wb, "User Workflows", headers, col_widths)
            
            # Populate with existing data
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    
            print(f"Populated User Workflows with {len(rows)} existing records")
            return ws
    
    # Create new worksheet if no existing data or headers don't match
    ws = setup_worksheet(wb, "User Workflows", headers, col_widths)
    
    # Add workflow steps
    workflows = [
        # Process Order workflow
        {
            "workflow": "Process Order",
            "actor": "Barista",
            "step": 1,
            "action": "View pending orders in queue",
            "component": "PendingOrdersSection",
            "route": "/api/orders/pending",
            "notes": "Orders are sorted by priority and wait time",
            "screenshot": "Image 9 (Upcoming Orders section)"
        },
        {
            "workflow": "Process Order",
            "actor": "Barista",
            "step": 2,
            "action": "Start preparing a specific order",
            "component": "PendingOrderCard",
            "route": "/api/orders/{id}/start",
            "notes": "Clicking Start moves order to in-progress",
            "screenshot": "Image 9 (Start buttons)"
        },
        {
            "workflow": "Process Order",
            "actor": "Barista",
            "step": 3,
            "action": "Prepare the coffee according to specifications",
            "component": "CurrentOrderSection",
            "route": "None",
            "notes": "Details shown include coffee type, milk, sugar, etc.",
            "screenshot": "Image 9 (Current Order section)"
        },
        {
            "workflow": "Process Order",
            "actor": "Barista",
            "step": 4,
            "action": "Mark order as completed",
            "component": "CurrentOrderSection",
            "route": "/api/orders/{id}/complete",
            "notes": "Sends automatic SMS notification to customer",
            "screenshot": "Image 9 (COMPLETE ORDER button)"
        },
        {
            "workflow": "Process Order",
            "actor": "Customer",
            "step": 5,
            "action": "Receive notification and view order status",
            "component": "DisplayScreen",
            "route": "/api/orders/completed",
            "notes": "SMS notification and display screen show completion",
            "screenshot": "Image 2 (Ready for Pickup section)"
        },
        
        # Batch Processing workflow
        {
            "workflow": "Batch Processing",
            "actor": "Barista",
            "step": 1,
            "action": "Identify similar orders in the batch groups section",
            "component": "BatchGroupsList",
            "route": "/api/orders/pending",
            "notes": "System automatically groups similar orders",
            "screenshot": "Image 9 (Batch Groups section)"
        },
        {
            "workflow": "Batch Processing",
            "actor": "Barista",
            "step": 2,
            "action": "Click Process Batch to start all orders in group",
            "component": "BatchGroupsList",
            "route": "/api/orders/batch",
            "notes": "Multiple orders moved to in-progress at once",
            "screenshot": "Image 9 (Process Batch button)"
        },
        {
            "workflow": "Batch Processing",
            "actor": "Barista",
            "step": 3,
            "action": "Prepare all coffees in the batch",
            "component": "CurrentOrderSection",
            "route": "None",
            "notes": "Efficient preparation of similar drinks",
            "screenshot": "N/A"
        },
        {
            "workflow": "Batch Processing",
            "actor": "Barista",
            "step": 4,
            "action": "Complete each order individually",
            "component": "CurrentOrderSection",
            "route": "/api/orders/{id}/complete",
            "notes": "Each customer gets individual notification",
            "screenshot": "Image 9 (COMPLETE ORDER button)"
        },
        
        # Customer Pickup workflow
        {
            "workflow": "Customer Pickup",
            "actor": "Customer",
            "step": 1,
            "action": "Receive SMS notification that order is ready",
            "component": "None (SMS Service)",
            "route": "None",
            "notes": "Automatic notification when barista marks complete",
            "screenshot": "N/A"
        },
        {
            "workflow": "Customer Pickup",
            "actor": "Customer",
            "step": 2,
            "action": "View order on display screen",
            "component": "DisplayScreen",
            "route": "/api/orders/completed",
            "notes": "Customer's name and order details displayed",
            "screenshot": "Image 2 (Ready for Pickup section)"
        },
        {
            "workflow": "Customer Pickup",
            "actor": "Customer",
            "step": 3,
            "action": "Collect order from specified station",
            "component": "None",
            "route": "None",
            "notes": "Station number shown in SMS and on display",
            "screenshot": "Image 2 (Station #1 - Main Hall)"
        },
        {
            "workflow": "Customer Pickup",
            "actor": "Barista",
            "step": 4,
            "action": "Mark order as picked up",
            "component": "CompletedOrderCard",
            "route": "/api/orders/{id}/pickup",
            "notes": "Removes order from display screen",
            "screenshot": "N/A"
        },
        
        # Inventory Management workflow
        {
            "workflow": "Inventory Management",
            "actor": "Barista",
            "step": 1,
            "action": "Check inventory levels in Stock tab",
            "component": "StockTab",
            "route": "/api/inventory",
            "notes": "View current stock levels for all items",
            "screenshot": "Image 14 (Station Inventory Status)"
        },
        {
            "workflow": "Inventory Management",
            "actor": "Barista",
            "step": 2,
            "action": "Update inventory when restocking",
            "component": "StockTab",
            "route": "/api/inventory/update",
            "notes": "Use + and - buttons to adjust stock levels",
            "screenshot": "Image 14 (Restock buttons)"
        },
        {
            "workflow": "Inventory Management",
            "actor": "System",
            "step": 3,
            "action": "Generate low stock alerts",
            "component": "None (Alert System)",
            "route": "None",
            "notes": "Automatic notification when items run low",
            "screenshot": "Image 8, Images 10-17 (Low Stock Alert)"
        },
        
        # System Monitoring workflow
        {
            "workflow": "System Monitoring",
            "actor": "Support Staff",
            "step": 1,
            "action": "Check overall system status",
            "component": "SystemMonitoring",
            "route": "/api/monitoring",
            "notes": "View health of all system components",
            "screenshot": "Image 6 (Service Status)"
        },
        {
            "workflow": "System Monitoring",
            "actor": "Support Staff",
            "step": 2,
            "action": "Review recent system logs",
            "component": "SystemLogs",
            "route": "/api/logs",
            "notes": "Check for errors and warnings",
            "screenshot": "Image 5 (System Logs)"
        },
        {
            "workflow": "System Monitoring",
            "actor": "Support Staff",
            "step": 3,
            "action": "Restart services if needed",
            "component": "SystemMonitoring",
            "route": "/api/services/{service}/restart",
            "notes": "Recover from degraded services",
            "screenshot": "Image 6 (Restart buttons)"
        }
    ]
    
    for row_idx, step in enumerate(workflows, 2):
        ws.cell(row=row_idx, column=1, value=step["workflow"])
        ws.cell(row=row_idx, column=2, value=step["actor"])
        ws.cell(row=row_idx, column=3, value=step["step"])
        ws.cell(row=row_idx, column=4, value=step["action"])
        ws.cell(row=row_idx, column=5, value=step["component"])
        ws.cell(row=row_idx, column=6, value=step["route"])
        ws.cell(row=row_idx, column=7, value=step["notes"])
        ws.cell(row=row_idx, column=8, value=step["screenshot"])
    
    # Group by workflow and apply consistent formatting
    current_workflow = None
    workflow_start_row = 2
    
    for row_idx in range(2, len(workflows) + 2):
        workflow = ws.cell(row=row_idx, column=1).value
        
        # Apply alternating row colors
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color=COLORS["alt_row"],
                    end_color=COLORS["alt_row"],
                    fill_type="solid"
                )
        
        # If we've moved to a new workflow, add workflow highlighting
        if workflow != current_workflow:
            if current_workflow is not None:
                # Apply formatting to the previous workflow section
                for r in range(workflow_start_row, row_idx):
                    ws.cell(row=r, column=1).font = Font(bold=(r == workflow_start_row))
            
            current_workflow = workflow
            workflow_start_row = row_idx
    
    # Handle the last workflow
    if current_workflow is not None:
        for r in range(workflow_start_row, len(workflows) + 2):
            ws.cell(row=r, column=1).font = Font(bold=(r == workflow_start_row))
    
    print(f"Created User Workflows worksheet with {len(workflows)} entries")
    return ws

def populate_authentication_flow(wb, existing_data=None):
    """Populate the Authentication Flow worksheet."""
    headers = [
        "Process Step", "Frontend Component", "Backend Route", 
        "Required Data", "Response Data", "Security Measures", "UI Representation"
    ]
    col_widths = [25, 25, 30, 30, 30, 40, 30]
    
    if existing_data and "Authentication Flow" in existing_data:
        rows = existing_data["Authentication Flow"]["rows"]
        # Verify headers match
        if existing_data["Authentication Flow"]["headers"] == headers:
            ws = setup_worksheet(wb, "Authentication Flow", headers, col_widths)
            
            # Populate with existing data
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    
            print(f"Populated Authentication Flow with {len(rows)} existing records")
            return ws
    
    # Create new worksheet if no existing data or headers don't match
    ws = setup_worksheet(wb, "Authentication Flow", headers, col_widths)
    
    # Add authentication flow steps
    auth_steps = [
        {
            "step": "User Login",
            "component": "LoginPage",
            "route": "/api/auth/login",
            "required": "username, password",
            "response": "JWT token, refresh token, user object with role",
            "security": "Password hashing, HTTPS, rate limiting",
            "ui": "Role selection screen (Image 18)"
        },
        {
            "step": "Role Selection",
            "component": "RoleSelectionPage",
            "route": "None (client-side)",
            "required": "User role from login response",
            "response": "None (client-side navigation)",
            "security": "Role validation, secure tokens in localStorage",
            "ui": "Role selection screen (Image 18)"
        },
        {
            "step": "Token Storage",
            "component": "AuthService.js",
            "route": "None (client-side)",
            "required": "JWT token, refresh token",
            "response": "None",
            "security": "Secure localStorage with token expiration",
            "ui": "No direct UI representation"
        },
        {
            "step": "Protected Route Access",
            "component": "ProtectedRoute",
            "route": "None (client-side)",
            "required": "Valid JWT token, user role",
            "response": "None (route access or redirect)",
            "security": "Role-based access control, token validation",
            "ui": "Redirects unauthenticated users to login"
        },
        {
            "step": "API Request with Auth",
            "component": "OrderDataService",
            "route": "All protected API endpoints",
            "required": "JWT token in Authorization header",
            "response": "Varies by endpoint",
            "security": "Server-side token validation",
            "ui": "No direct UI representation"
        },
        {
            "step": "Token Refresh",
            "component": "AuthService.refreshToken()",
            "route": "/api/auth/refresh",
            "required": "Valid refresh token",
            "response": "New JWT token",
            "security": "Refresh token rotation, secure storage",
            "ui": "Happens automatically when token expires"
        },
        {
            "step": "User Logout",
            "component": "Logout button",
            "route": "/api/auth/logout",
            "required": "None",
            "response": "Success status",
            "security": "Token invalidation, clear localStorage",
            "ui": "Logout button in navigation (not visible in screenshots)"
        },
        {
            "step": "Session Timeout",
            "component": "AuthService.handleTimeout()",
            "route": "None (client-side)",
            "required": "None",
            "response": "None",
            "security": "Automatic logout on token expiration",
            "ui": "Session timeout dialog (not visible in screenshots)"
        },
        {
            "step": "Access Control",
            "component": "ProtectedComponents",
            "route": "Backend middleware",
            "required": "Valid JWT token with appropriate role",
            "response": "403 Forbidden if unauthorized",
            "security": "Role-based access control, middleware guards",
            "ui": "Different interfaces based on role (Image 18)"
        },
        {
            "step": "Special Access Codes",
            "component": "Access code inputs",
            "route": "/api/auth/validate-code",
            "required": "VIP code, staff code, or display code",
            "response": "Valid status and associated permissions",
            "security": "Unique codes, server validation",
            "ui": "Access code fields in Image 11 (System Settings)"
        }
    ]
    
    for row_idx, step in enumerate(auth_steps, 2):
        ws.cell(row=row_idx, column=1, value=step["step"])
        ws.cell(row=row_idx, column=2, value=step["component"])
        ws.cell(row=row_idx, column=3, value=step["route"])
        ws.cell(row=row_idx, column=4, value=step["required"])
        ws.cell(row=row_idx, column=5, value=step["response"])
        ws.cell(row=row_idx, column=6, value=step["security"])
        ws.cell(row=row_idx, column=7, value=step["ui"])
    
    # Apply alternating row colors
    for row_idx in range(2, len(auth_steps) + 2):
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color=COLORS["alt_row"],
                    end_color=COLORS["alt_row"],
                    fill_type="solid"
                )
    
    print(f"Created Authentication Flow worksheet with {len(auth_steps)} entries")
    return ws

def populate_ui_components(wb, existing_data=None):
    """Populate the UI Components worksheet with information from screenshots."""
    headers = [
        "Screen Name", "Purpose", "Components", "Routes/Endpoints",
        "User Roles", "Key Features", "Navigation Path", "Screenshot Reference"
    ]
    col_widths = [25, 40, 35, 35, 15, 40, 25, 20]
    
    if existing_data and "UI Components" in existing_data:
        rows = existing_data["UI Components"]["rows"]
        # Verify headers match
        if existing_data["UI Components"]["headers"] == headers:
            ws = setup_worksheet(wb, "UI Components", headers, col_widths)
            
            # Populate with existing data
            for row_idx, row_data in enumerate(rows, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    
            print(f"Populated UI Components with {len(rows)} existing records")
            return ws
    
    # Create new worksheet if no existing data or headers don't match
    ws = setup_worksheet(wb, "UI Components", headers, col_widths)
    
    # Add UI screens from screenshots
    screens = [
        {
            "name": "Barista Interface",
            "purpose": "Main workspace for baristas to process orders",
            "components": "HeaderBar, NavTabs, CurrentOrderSection, PendingOrdersSection, ActionBar",
            "routes": "/api/orders/pending, /api/orders/in-progress, /api/orders/{id}/start, /api/orders/{id}/complete",
            "roles": "Barista",
            "features": "Order processing, batch handling, wait time management, stock monitoring",
            "navigation": "Role Selection → Barista → Orders tab",
            "screenshot": "Image 9"
        },
        {
            "name": "Customer Display Screen",
            "purpose": "Public-facing screen showing order status to customers",
            "components": "InProgressList, ReadyForPickupList, WaitTimeDisplay",
            "routes": "/api/orders/in-progress, /api/orders/completed",
            "roles": "Display",
            "features": "Real-time order status, preparation indicators, pickup notifications",
            "navigation": "Role Selection → Display Screen",
            "screenshot": "Image 2"
        },
        {
            "name": "System Settings",
            "purpose": "Configure global system settings and parameters",
            "components": "SettingsForm, NotificationSettings, SecuritySettings",
            "routes": "/api/settings",
            "roles": "Admin",
            "features": "Event details, wait time defaults, SMS configuration, access codes",
            "navigation": "Role Selection → Admin → Settings",
            "screenshot": "Image 3, Image 10, Image 11"
        },
        {
            "name": "Station Management",
            "purpose": "Monitor and configure coffee stations",
            "components": "StationList, StationStatus, StationConfig",
            "routes": "/api/stations, /api/stations/{id}",
            "roles": "Admin, Staff",
            "features": "Barista assignment, status updates, maintenance controls",
            "navigation": "Role Selection → Admin → Stations",
            "screenshot": "Image 4, Image 14"
        },
        {
            "name": "System Logs",
            "purpose": "Monitor system activity and troubleshoot issues",
            "components": "LogViewer, LogFilterControl",
            "routes": "/api/logs",
            "roles": "Support",
            "features": "Filterable log entries, error highlighting, service attribution",
            "navigation": "Role Selection → Support → Logs",
            "screenshot": "Image 5"
        },
        {
            "name": "System Monitoring",
            "purpose": "Monitor health and performance of system components",
            "components": "ServiceStatus, MonitoringOverview, ServiceControls",
            "routes": "/api/monitoring, /api/services/{service}/restart",
            "roles": "Support",
            "features": "Health indicators, restart controls, performance metrics",
            "navigation": "Role Selection → Support → Monitoring",
            "screenshot": "Image 6, Image 7"
        },
        {
            "name": "Support Dashboard",
            "purpose": "Overview of system health and activity for support staff",
            "components": "StatusCards, ActiveAlerts, RecentLogs",
            "routes": "/api/stats, /api/alerts, /api/logs",
            "roles": "Support",
            "features": "System metrics, active issue tracking, quick service restart",
            "navigation": "Role Selection → Support → Dashboard",
            "screenshot": "Image 8"
        },
        {
            "name": "Messaging & Notifications",
            "purpose": "Manage SMS templates and send batch messages",
            "components": "MessageTemplates, RecipientSelector, MessageStats",
            "routes": "/api/messaging/templates, /api/messaging/send-batch",
            "roles": "Admin, Staff",
            "features": "Template management, batch messaging, delivery statistics",
            "navigation": "Role Selection → Admin → Messaging",
            "screenshot": "Image 12"
        },
        {
            "name": "Event Schedule",
            "purpose": "Manage conference schedule and coffee breaks",
            "components": "ScheduleCalendar, BreakManagement, SessionList",
            "routes": "/api/schedule, /api/breaks",
            "roles": "Admin, Staff",
            "features": "Session scheduling, break management, pre-order coordination",
            "navigation": "Role Selection → Admin → Schedule",
            "screenshot": "Image 13"
        },
        {
            "name": "Analytics Dashboard",
            "purpose": "Visualize order data and system metrics",
            "components": "OrderChart, CoffeeTypeChart, CustomerStats",
            "routes": "/api/analytics/orders, /api/analytics/customers",
            "roles": "Admin, Staff",
            "features": "Volume trends, preference analytics, peak time analysis",
            "navigation": "Role Selection → Admin → Analytics",
            "screenshot": "Image 15, Image 17"
        },
        {
            "name": "Order Management",
            "purpose": "Monitor and manage all orders in the system",
            "components": "OrderTable, OrderFilters, ActionButtons",
            "routes": "/api/orders, /api/orders/{id}",
            "roles": "Admin, Staff",
            "features": "Order reassignment, status updates, batch processing",
            "navigation": "Role Selection → Admin → Orders",
            "screenshot": "Image 16"
        },
        {
            "name": "Role Selection",
            "purpose": "Entry point to select user role and access appropriate interface",
            "components": "RoleCards, LoginButtons",
            "routes": "/api/auth/login",
            "roles": "All",
            "features": "Role-based access control, authentication initiation",
            "navigation": "Login → Role Selection",
            "screenshot": "Image 18"
        }
    ]
    
    for row_idx, screen in enumerate(screens, 2):
        ws.cell(row=row_idx, column=1, value=screen["name"])
        ws.cell(row=row_idx, column=2, value=screen["purpose"])
        ws.cell(row=row_idx, column=3, value=screen["components"])
        ws.cell(row=row_idx, column=4, value=screen["routes"])
        ws.cell(row=row_idx, column=5, value=screen["roles"])
        ws.cell(row=row_idx, column=6, value=screen["features"])
        ws.cell(row=row_idx, column=7, value=screen["navigation"])
        ws.cell(row=row_idx, column=8, value=screen["screenshot"])
    
    # Apply alternating row colors
    for row_idx in range(2, len(screens) + 2):
        if row_idx % 2 == 0:
            for col_idx in range(1, len(headers) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = PatternFill(
                    start_color=COLORS["alt_row"],
                    end_color=COLORS["alt_row"],
                    fill_type="solid"
                )
    
    print(f"Created UI Components worksheet with {len(screens)} entries")
    return ws

def populate_readme(wb):
    """Create a README worksheet with information about the documentation."""
    ws = wb.create_sheet("README", 0)
    
    # Set column widths
    ws.column_dimensions['A'].width = 100
    
    # Add title
    ws.merge_cells('A1:A2')
    cell = ws.cell(row=1, column=1, value="Coffee Cue System Documentation")
    cell.font = Font(size=16, bold=True, color="000000")
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    
    # Add introduction
    row = 4
    ws.cell(row=row, column=1, value="Introduction").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="This Excel workbook contains comprehensive documentation for the Coffee Cue System, " +
                                   "a coffee ordering and management application for events.")
    row += 2
    
    # Add worksheets description
    ws.cell(row=row, column=1, value="Worksheets:").font = Font(bold=True, size=12)
    row += 1
    
    worksheets = [
        ("File Components", "Lists all system files with their purpose and dependencies"),
        ("Frontend Architecture", "Details React components and their relationships"),
        ("Backend Architecture", "Documents Flask routes, handlers, and their requirements"),
        ("Database Schema", "Describes database tables, columns, and relationships"),
        ("API Endpoints", "Lists all API endpoints with request/response formats"),
        ("Authentication Flow", "Explains the authentication process steps"),
        ("User Workflows", "Describes step-by-step user scenarios"),
        ("UI Components", "Maps UI screens from screenshots to components and functionality")
    ]
    
    for sheet_name, description in worksheets:
        ws.cell(row=row, column=1, value=f"• {sheet_name}: {description}")
        row += 1
    
    row += 2
    
    # Add usage information
    ws.cell(row=row, column=1, value="How to Use This Documentation").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="• Each worksheet contains different aspects of the system")
    row += 1
    ws.cell(row=row, column=1, value="• Use filters and sorting to find relevant information")
    row += 1
    ws.cell(row=row, column=1, value="• Cross-reference between worksheets for complete understanding")
    row += 1
    ws.cell(row=row, column=1, value="• Screenshot references refer to the numbered images in the documentation folder")
    row += 1
    
    row += 2
    
    # Add notes for AI assistants
    ws.cell(row=row, column=1, value="Notes for AI Assistants").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="• This documentation is designed to help both humans and AI assistants understand the system")
    row += 1
    ws.cell(row=row, column=1, value="• Key relationships are documented to provide context for code modifications")
    row += 1
    ws.cell(row=row, column=1, value="• UI screenshots have been analyzed and mapped to code components")
    row += 1
    ws.cell(row=row, column=1, value="• When suggesting changes, reference this documentation to maintain consistency")
    row += 1
    
    row += 2
    
    # Add maintenance instructions
    ws.cell(row=row, column=1, value="Maintenance").font = Font(bold=True, size=12)
    row += 1
    ws.cell(row=row, column=1, value="• Update this documentation when making significant changes to the system")
    row += 1
    ws.cell(row=row, column=1, value="• Use the documentation generator script to maintain consistency")
    row += 1
    ws.cell(row=row, column=1, value=f"• Last updated: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    row += 1
    
    # Set row heights
    for i in range(1, row):
        ws.row_dimensions[i].height = 20
    
    print("Created README worksheet")
    return ws

def generate_documentation():
    """Generate the complete Excel documentation."""
    # Load existing data if available
    existing_data = load_existing_documentation()
    
    # Create a new workbook
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])  # Remove default sheet
    
    # Create README sheet
    populate_readme(wb)
    
    # Create all documentation worksheets
    populate_file_components(wb, existing_data)
    populate_frontend_architecture(wb, existing_data)
    populate_backend_architecture(wb, existing_data)
    populate_database_schema(wb, existing_data)
    populate_api_endpoints(wb, existing_data)
    populate_authentication_flow(wb, existing_data)
    populate_user_workflows(wb, existing_data)
    populate_ui_components(wb, existing_data)
    
    # Save the workbook
    try:
        wb.save(OUTPUT_FILE)
        print(f"Documentation saved to {OUTPUT_FILE}")
        return True
    except Exception as e:
        print(f"Error saving documentation: {e}")
        return False

if __name__ == "__main__":
    print("Coffee Cue System Documentation Generator")
    print("----------------------------------------")
    
    success = generate_documentation()
    
    if success:
        print("\nDocumentation generation complete!")
        print(f"Excel file saved to: {os.path.abspath(OUTPUT_FILE)}")
    else:
        print("\nDocumentation generation failed.")
        sys.exit(1)
